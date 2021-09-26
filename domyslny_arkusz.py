from datetime import *
from openpyxl import *
from openpyxl.styles import Font, colors, PatternFill
from openpyxl.styles import Color

def days_cur_month():
    """Listing all days in curent month"""
    m = datetime.now().month
    y = datetime.now().year
    ndays = (date(y, m+1, 1) - date(y, m, 1)).days
    d1 = date(y, m, 1)
    d2 = date(y, m, ndays)
    delta = d2 - d1
    list_days = [(d1 + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(delta.days + 1)]
    return list_days


def blank_sheet_create(name='Adam Wawrzyniak'):
    """Creating a blank sheet with default text and making a good looking
    dimensions of columns"""
    wb = Workbook()
    sheet = wb.active
    sheet["A1"] = name
    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 10
    sheet.column_dimensions['E'].width = 18
    sheet.merge_cells('A1:B1')
    sheet.merge_cells('B2:C2')
    sheet.merge_cells('C1:F1')
    sheet["A2"] = datetime.today().strftime("%m-%y")
    sheet["B2"] = 'suma godzin'
    sheet["E2"] = 'Uwagi/lokalizacja'

    """There we list all month dates in rows"""
    days_cur_month()
    n = 3
    for value in days_cur_month():
        sheet.cell(row=n, column=1).value = value
        d = datetime(int(value[:4]), int(value[5:7]), int(value[8:10]))
        if d.weekday() > 4:
            for col in range(2 , 5):
                weekend = sheet.cell(row=n, column=col)
                weekend.fill = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")
            n += 1
        else:
            nulltime = "00:00:00"
            for col in range(2, 4):
                zero = sheet.cell(row=n, column=col)
                zero.value = nulltime
                # Do poprawy... w komórkach wbija się pojedyńczy cudzysłów
                # przed nulltime. To samo dzieje się przy listowaniu daty.
            n += 1
            continue
    """Sum hours """
    last_row = sheet.max_row

    for i in range(3, last_row+1):
        sheet.cell(row=i, column=4).value = f"=SUM(B{i}:C{i})"
        wb.save(f'Godziny{datetime.today().strftime("%m-%y")}.xlsx')